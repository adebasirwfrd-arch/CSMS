#!/usr/bin/env python3
"""Import PTS Wells Matrix Excel into data/matrix_workbook.json (+ Supabase seed)."""
import json
import re
import sys
import uuid
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

import openpyxl
DEFAULT_EXCEL = Path.home() / "Downloads" / "PTS Wells Block A (1).xlsx"
OUT = ROOT / "data" / "matrix_workbook.json"


def slug(s: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", (s or "").lower().strip())
    return s.strip("_")[:48] or "sheet"


def serialize(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def col_type(label: str) -> str:
    l = label.lower()
    if "date" in l or "expiry" in l or "expired" in l:
        return "date"
    if l in ("no",) or l.startswith("no "):
        return "number"
    if "email" in l:
        return "email"
    if "phone" in l or "whatsapp" in l or " wa" in l or "hp number" in l:
        return "phone"
    if "gender" in l:
        return "select"
    if any(x in l for x in ("result", "status", "vaccine", "education", "relation", "family status")):
        return "select"
    return "text"


def is_filterable(label: str, ctype: str, samples: list) -> bool:
    l = label.lower()
    if ctype == "date":
        return True
    if any(
        k in l
        for k in (
            "gender",
            "personnel name",
            "company name",
            "client name",
            "home base",
            "working location",
            "budget type",
            "result",
            "status",
            "vaccine",
            "country",
            "city",
            "education",
            "contract name",
        )
    ):
        return True
    uniq = {s for s in samples if s}
    return 0 < len(uniq) <= 30


def import_workbook(excel_path: Path) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    workbook = {"version": 1, "updated_at": datetime.utcnow().isoformat() + "Z", "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        title = serialize(ws.cell(1, 1).value) or sheet_name.strip()
        header_row = 2
        cols_meta = []
        for c in range(1, ws.max_column + 1):
            label = ws.cell(header_row, c).value
            if label is None or not str(label).strip():
                continue
            label = str(label).strip()
            cols_meta.append({"index": c, "id": f"col_{c}", "label": label})

        samples_by_col = {m["id"]: [] for m in cols_meta}
        rows_out = []
        for r in range(header_row + 1, ws.max_row + 1):
            cells = {}
            has_data = False
            for m in cols_meta:
                v = serialize(ws.cell(r, m["index"]).value)
                cells[m["id"]] = v
                if v:
                    has_data = True
                if len(samples_by_col[m["id"]]) < 200:
                    samples_by_col[m["id"]].append(v)
            if has_data:
                rows_out.append({"id": f"row_{uuid.uuid4().hex[:12]}", "cells": cells})

        columns = []
        for m in cols_meta:
            ctype = col_type(m["label"])
            columns.append(
                {
                    "id": m["id"],
                    "key": slug(m["label"]) + f"_{m['index']}",
                    "label": m["label"],
                    "type": ctype,
                    "filterable": is_filterable(m["label"], ctype, samples_by_col[m["id"]]),
                    "required": "*" in m["label"],
                    "index": m["index"],
                }
            )

        workbook["sheets"].append(
            {
                "id": slug(sheet_name),
                "name": sheet_name.strip(),
                "title": title,
                "columns": columns,
                "rows": rows_out,
            }
        )
    return workbook


def main():
    excel = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not excel.exists():
        print(f"File not found: {excel}")
        sys.exit(1)
    data = import_workbook(excel)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    for s in data["sheets"]:
        print(f"  {s['id']}: {len(s['columns'])} cols, {len(s['rows'])} rows")
    print(f"Saved JSON: {OUT}")
    try:
        from services.matrix_store import SUPABASE_MATRIX, seed_workbook

        if not SUPABASE_MATRIX:
            print("Supabase not configured — JSON only.")
            print("Set SUPABASE_URL + SUPABASE_KEY in .env lalu jalankan ulang script ini.")
            return
        seed_workbook(data)
        print("OK — Seeded Supabase: matrix_sheets, matrix_columns, matrix_rows")
    except Exception as e:
        print(f"Supabase seed failed: {e}")
        print("Pastikan create_matrix_tables.sql sudah di-run di SQL Editor, lalu coba lagi.")
        sys.exit(1)


if __name__ == "__main__":
    main()
