"""Import / sync product line employees from Excel or bundled seed JSON."""
from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from services.product_line_employee_utils import normalize_yes_no

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = ROOT / "Blueprint" / "Employee_Job_Data_2026.xlsx"
SEED_JSON = ROOT / "data" / "product_line_employees_seed.json"

# Excel header → internal key
_EXCEL_HEADERS = {
    "NO": "row_no",
    "NAME": "name",
    "Job Family Description": "job_family_description",
    "Job Description": "job_description",
    "ACCESS TO PL": "access_to_pl",
    "ACCESS PERSONNEL ONLY": "access_personnel_only",
    "Email": "email",
    "Product Line": "product_line",
}


def _norm_name(value: str) -> str:
    return (value or "").strip().lower()


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _parse_row_no(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def load_seed_payload(path: Optional[Path] = None) -> Dict[str, Any]:
    seed_path = path or SEED_JSON
    with open(seed_path, encoding="utf-8") as f:
        return json.load(f)


def load_from_excel(excel_path: Optional[Path] = None) -> Dict[str, Any]:
    xlsx = excel_path or DEFAULT_EXCEL
    if not xlsx.exists():
        raise FileNotFoundError(f"Excel not found: {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError("Excel sheet is empty")

        col_map: Dict[int, str] = {}
        for idx, label in enumerate(header_row):
            key = _EXCEL_HEADERS.get(_clean_cell(label), "")
            if key:
                col_map[idx] = key

        employees: List[Dict[str, Any]] = []
        for row in rows_iter:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            record: Dict[str, Any] = {
                "product_line": "",
                "row_no": None,
                "name": "",
                "job_family_description": "",
                "job_description": "",
                "access_to_pl": "",
                "access_personnel_only": "",
                "email": "",
            }
            for idx, field in col_map.items():
                val = row[idx] if idx < len(row) else None
                if field == "row_no":
                    record["row_no"] = _parse_row_no(val)
                else:
                    record[field] = _clean_cell(val)
            if record.get("name") or record.get("product_line"):
                employees.append(record)
    finally:
        wb.close()

    product_lines = sorted({e["product_line"] for e in employees if e["product_line"]})
    try:
        source = str(xlsx.relative_to(ROOT))
    except ValueError:
        source = str(xlsx)
    return {
        "source": source,
        "product_lines": product_lines,
        "employees": employees,
    }


def load_import_payload(
    *,
    use_excel: bool = False,
    excel_path: Optional[Path] = None,
    seed_path: Optional[Path] = None,
) -> Dict[str, Any]:
    if use_excel:
        try:
            return load_from_excel(excel_path)
        except (FileNotFoundError, ImportError, ValueError):
            pass
    return load_seed_payload(seed_path)


def resolve_product_line_id(
    excel_name: str, product_lines: List[Dict[str, Any]]
) -> Optional[int]:
    key = _norm_name(excel_name)
    if not key:
        return None
    for pl in product_lines:
        if _norm_name(pl.get("name", "")) == key:
            return pl.get("id")
    return None


def sync_product_lines_and_employees(
    payload: Dict[str, Any],
    *,
    create_product_line,
    get_product_lines,
    replace_employees_for_product_line,
) -> Dict[str, Any]:
    """
    Ensure product lines from payload exist, then replace employee rows per line.
    """
    existing = get_product_lines()
    by_norm = {_norm_name(p.get("name", "")): p for p in existing}
    created_lines: List[str] = []
    updated_counts: Dict[str, int] = {}
    skipped: List[str] = []

    for pl_name in payload.get("product_lines") or []:
        key = _norm_name(pl_name)
        if not key:
            continue
        if key not in by_norm:
            created = create_product_line({"name": pl_name, "description": None})
            by_norm[key] = created
            created_lines.append(pl_name)

    all_lines = get_product_lines()
    employees = payload.get("employees") or []
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for emp in employees:
        pl_name = emp.get("product_line") or ""
        grouped.setdefault(pl_name, []).append(emp)

    for pl_name, rows in grouped.items():
        pl_id = resolve_product_line_id(pl_name, all_lines)
        if not pl_id:
            skipped.append(pl_name)
            continue
        sorted_rows = sorted(
            rows,
            key=lambda r: (r.get("row_no") is None, r.get("row_no") or 0, r.get("name") or ""),
        )
        records = [
            {
                "product_line_id": pl_id,
                "row_no": r.get("row_no"),
                "name": r.get("name") or "",
                "job_family_description": r.get("job_family_description") or "",
                "job_description": r.get("job_description") or "",
                "access_to_pl": normalize_yes_no(r.get("access_to_pl")),
                "access_personnel_only": normalize_yes_no(r.get("access_personnel_only")),
                "email": r.get("email") or "",
                "email_reminder": normalize_yes_no(r.get("email_reminder")),
            }
            for r in sorted_rows
        ]
        replace_employees_for_product_line(pl_id, records)
        updated_counts[pl_name] = len(records)

    return {
        "source": payload.get("source"),
        "product_lines_created": created_lines,
        "employees_by_product_line": updated_counts,
        "skipped_product_lines": skipped,
        "total_employees": sum(updated_counts.values()),
    }
